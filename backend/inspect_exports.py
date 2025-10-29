import sys
from pathlib import Path
import openpyxl

exports_dir = Path(__file__).resolve().parent / 'exports'
if not exports_dir.exists():
    print('exports directory not found:', exports_dir)
    sys.exit(1)

files = sorted(exports_dir.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
if not files:
    print('No files found in exports directory:', exports_dir)
    sys.exit(1)

latest = files[0]
print('Inspecting latest export:', latest)

try:
    wb = openpyxl.load_workbook(str(latest))
except Exception as e:
    print('Failed to open workbook:', e)
    sys.exit(2)

for ws in wb.worksheets:
    print('\nSheet:', ws.title)
    max_row = min(ws.max_row, 20)
    max_col = min(ws.max_column, 40)
    print('showing rows 2..6 and cols 3..10 (if present)')
    for r in range(2, min(7, max_row+1)):
        row_out = []
        for c in range(3, min(11, max_col+1)):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            fill = cell.fill
            pat = getattr(fill, 'patternType', None)
            start = getattr(getattr(fill, 'start_color', None), 'rgb', None) or getattr(getattr(fill, 'start_color', None), 'index', None)
            row_out.append(f"({r},{c}) val={repr(val)} fill_pattern={pat} fill_color={start}")
        print(' | '.join(row_out))

# Also quick summary: count X cells and how many have a fill set
print('\nSummary across all sheets:')
for ws in wb.worksheets:
    x_count = 0
    filled = 0
    total = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            total += 1
            if any(ch in s for ch in ('x','X','✕','✖','×','✘')):
                x_count += 1
                pat = getattr(cell.fill, 'patternType', None)
                if pat:
                    filled += 1
    print(f"Sheet {ws.title}: total non-empty={total}, X-like={x_count}, X-with-fill={filled}")

print('\nDone')
