from pathlib import Path
import openpyxl

exports_dir = Path(__file__).resolve().parent / 'exports'
files = sorted(exports_dir.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
if not files:
    print('No exports')
    raise SystemExit(1)
latest = files[0]
print('Latest:', latest)
wb = openpyxl.load_workbook(str(latest))
for ws in wb.worksheets:
    print('\nSheet', ws.title)
    found = []
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            if any(ch in s for ch in ('x','X','✕','✖','×','✘')):
                pat = getattr(cell.fill, 'patternType', None)
                if pat:
                    found.append((r,c,s,pat, getattr(cell.fill.start_color,'rgb',None) or getattr(cell.fill.start_color,'index',None)))
    print('Total X-with-fill:', len(found))
    for i,entry in enumerate(found[:200]):
        r,c,s,pat,col = entry
        print(f"{i+1}: {ws.title}!{openpyxl.utils.get_column_letter(c)}{r} = {s} fill={pat} color={col}")
