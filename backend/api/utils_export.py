import io
from datetime import datetime
from typing import Dict, Set

from django.conf import settings
from openpyxl import load_workbook

from .models import Registration, Attendance


def parse_local_date_from_ymd(date_str: str) -> datetime:
    parts = (date_str or "").split("-")
    try:
        year = int(parts[0])
        month = int(parts[1])
        day = int(parts[2])
        return datetime(year, month, day)
    except Exception:
        return datetime.utcnow()


def build_attendance_workbook(include_names: bool = False) -> io.BytesIO:
    """Build an Excel workbook (in-memory) based on the stored registrations and attendance.

    This function expects an Excel template file named 'attendance_template.xlsx' located
    in the Django static files or the project root (adjust path as necessary). It will
    mark 'X' for absent days and leave cells empty for present days consistent with the
    front-end logic.
    """
    # Load registrations and attendance records
    registrations = list(Registration.objects.all().order_by('student'))
    attendances = list(Attendance.objects.all())

    # Map attendance by normalized key -> month -> set(days)
    month_names = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

    def normalize_name(s: str) -> str:
        return (s or "").strip().lower().replace('\n', ' ').replace('\r', '').replace('\t', ' ')

    # Build mapping from registration LRN or name to dates
    attendance_by_key: Dict[str, Dict[str, Set[int]]] = {}
    for a in attendances:
        dt = a.time
        month = month_names[dt.month - 1]
        day = dt.day
        # keys: lrn, normalized name
        lrn = (a.student.lrn or '').strip()
        name = (a.student.student or '').strip()
        keys = [lrn, normalize_name(name)]
        for k in keys:
            if not k:
                continue
            attendance_by_key.setdefault(k, {}).setdefault(month, set()).add(day)

    # Determine template path: prefer Django STATIC_ROOT or project root
    # Try common locations
    import os

    possible = [
        os.path.join(settings.BASE_DIR, 'attendance_template.xlsx'),
        os.path.join(settings.BASE_DIR, 'static', 'attendance_template.xlsx'),
        os.path.join(settings.BASE_DIR, 'backend', 'static', 'attendance_template.xlsx'),
    ]
    template_path = None
    for p in possible:
        if os.path.exists(p):
            template_path = p
            break
    if not template_path:
        raise FileNotFoundError('attendance_template.xlsx not found in project; place it in project root or static folder')

    wb = load_workbook(template_path)

    for month in month_names:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]

        # Helper to read header row for days (similar to front-end getCellDay)
        def get_cell_day(cell):
            v = cell.value
            if v is None:
                return None
            if isinstance(v, int):
                return v
            if isinstance(v, str) and v.strip().isdigit():
                return int(v.strip())
            if isinstance(v, datetime):
                return v.day
            return None

        # Read date columns from row 10 (1-based)
        date_row = ws[10]
        date_cols = []
        for idx, cell in enumerate(date_row, start=1):
            d = get_cell_day(cell)
            if d is not None:
                date_cols.append((idx, d))

        # Male rows start 13, female at 64 as front-end expects
        males = [r for r in registrations if (r.sex or '').lower() == 'male']
        females = [r for r in registrations if (r.sex or '').lower() == 'female']

        # Helper fill rows
        def process_rows(start_row, regs):
            row = start_row
            for reg in regs:
                visible = (reg.student and reg.student.strip()) or (reg.lrn and reg.lrn.strip()) or 'Unknown Student'
                if include_names:
                    ws.cell(row=row, column=2).value = visible
                name_key = normalize_name(visible)
                lrn_key = (reg.lrn or '').strip()
                # collect present days
                present = set()
                if lrn_key in attendance_by_key and month in attendance_by_key[lrn_key]:
                    present.update(attendance_by_key[lrn_key][month])
                if name_key in attendance_by_key and month in attendance_by_key[name_key]:
                    present.update(attendance_by_key[name_key][month])

                for col_idx, day in date_cols:
                    cell = ws.cell(row=row, column=col_idx)
                    # marked present in attendance -> leave blank, else X
                    if day in present and day != 1:
                        cell.value = ''
                    else:
                        cell.value = 'X'
                row += 1

        process_rows(13, males)
        process_rows(64, females)

    # Save workbook to BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
