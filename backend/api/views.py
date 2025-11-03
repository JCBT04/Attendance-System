from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from django.utils import timezone
from django.conf import settings
import os
from .models import Registration, Attendance, DroppedRegistration
from .serializers import RegistrationSerializer, AttendanceSerializer, DroppedRegistrationSerializer
from rest_framework.decorators import permission_classes
from rest_framework.permissions import AllowAny
import io
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except Exception:
    openpyxl = None

from django.http import FileResponse

class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = Registration.objects.all().order_by('student')
    serializer_class = RegistrationSerializer


@api_view(['GET', 'POST'])
def record_attendance(request):
    """Record attendance via QR scan (POST) or list attendances (GET).

    GET: returns serialized list of Attendance records (same shape as the viewset list).
    POST: existing behavior — create an Attendance from provided 'lrn'.
    """
    if request.method == 'GET':
        # Return all attendance records (keeps compatibility with frontend fallback)
        records = Attendance.objects.all()
        serializer = AttendanceSerializer(records, many=True)
        return Response(serializer.data)

    # POST: record a new attendance
    data = request.data
    lrn = data.get('lrn')
    if not lrn:
        return Response({'error': 'LRN is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student = Registration.objects.get(lrn=lrn)
    except Registration.DoesNotExist:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

    Attendance.objects.create(student=student, time=timezone.now())
    return Response({'message': f'Attendance recorded for {student.student}'}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def attendance_today(request):
    """Fetch today’s attendance"""
    today = timezone.localdate()
    records = Attendance.objects.filter(time__date=today)
    serializer = AttendanceSerializer(records, many=True)
    return Response(serializer.data)

@api_view(['DELETE'])
def clear_all_registrations(request):
    Registration.objects.all().delete()
    return Response({"message": "All registrations deleted."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['DELETE'])
def clear_attendance(request):
    """Clear attendance records. Query param 'scope' accepts 'today' or 'all'. Defaults to 'today'."""
    scope = request.GET.get('scope', 'today')
    try:
        if scope == 'all':
            Attendance.objects.all().delete()
            return Response({"message": "All attendance records deleted."}, status=status.HTTP_204_NO_CONTENT)
        else:
            today = timezone.localdate()
            Attendance.objects.filter(time__date=today).delete()
            return Response({"message": "Today's attendance deleted."}, status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

    @action(detail=False, methods=['get'])
    def today(self, request):
        today_date = now().date()
        attendances = Attendance.objects.filter(timestamp__date=today_date)
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)


@api_view(['POST'])
def upload_excel(request):
    """Accept an uploaded Excel file and save it to the backend 'exports' folder."""
    f = request.FILES.get('file')
    if not f:
        return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

    # Ensure exports directory exists under BASE_DIR
    out_dir = settings.BASE_DIR / 'exports'
    try:
        os.makedirs(out_dir, exist_ok=True)
    except Exception:
        # In some environments BASE_DIR may be a string
        out_dir = os.path.join(str(settings.BASE_DIR), 'exports')
        os.makedirs(out_dir, exist_ok=True)

    filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{f.name}"
    dest_path = out_dir / filename if hasattr(out_dir, '__truediv__') else os.path.join(out_dir, filename)

    try:
        # Write uploaded chunks to disk
        with open(dest_path, 'wb') as dest:
            for chunk in f.chunks():
                dest.write(chunk)
    except Exception as e:
        return Response({'error': f'Failed to save file: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Provide a URL so clients can download the saved file. During development we expose
    # the 'exports' folder via a static serve URL defined in backend/urls.py
    download_url = f"/exports/{filename}"
    return Response({'message': 'Saved', 'filename': filename, 'url': download_url}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def generate_excel_export(request):
    """Generate an Excel export server-side and save it to exports/. Returns download URL."""
    if openpyxl is None:
        return Response({'error': 'openpyxl not installed on server'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    regs = Registration.objects.all().order_by('student')
    attends = Attendance.objects.all()

    # Attempt to load an Excel template (if present in repo root `public/attendance_template.xlsx`)
    today = timezone.localdate()
    month_name = today.strftime('%b').upper()
    template_path = None
    try:
        # public folder is assumed to live one level above backend/ (project root)
        template_path = settings.BASE_DIR.parent / 'public' / 'attendance_template.xlsx'
        if template_path.exists():
            wb = openpyxl.load_workbook(str(template_path))
            # use the month sheet if present, otherwise create/use active and set title
            if month_name in wb.sheetnames:
                ws = wb[month_name]
            else:
                ws = wb.active
                try:
                    ws.title = month_name
                except Exception:
                    # if title clashes, create a new sheet
                    ws = wb.create_sheet(title=month_name)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = month_name
    except Exception:
        # fallback to a fresh workbook if loading template fails
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = month_name

    # If the worksheet is newly created or template missing header, ensure header row exists
    try:
        # Header row: Day numbers 1..31 starting at column 3 if not already present
        existing_header = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if not existing_header or len(existing_header) < 3:
            ws.cell(row=1, column=1, value='Student')
            for d in range(1, 32):
                ws.cell(row=1, column=2 + d, value=d)
    except Exception:
        # best-effort: ensure header
        ws.cell(row=1, column=1, value='Student')
        for d in range(1, 32):
            ws.cell(row=1, column=2 + d, value=d)

    # Build attendance flags map keyed by lrn or student name lowercased
    # Flags: AM=1, PM=2 (bitmask)
    AM_FLAG = 1
    PM_FLAG = 2
    att_flags = {}
    for a in attends:
        try:
            dt = a.time
            d = dt.date()
            hour = dt.hour
        except Exception:
            continue
        if d.month != today.month or d.year != today.year:
            continue
        key = (a.student.lrn or a.student.student).strip().lower()
        daynum = d.day
        flag = AM_FLAG if hour < 12 else PM_FLAG
        att_flags.setdefault(key, {}).setdefault(daynum, 0)
        att_flags[key][daynum] = att_flags[key][daynum] | flag

    # Render rows and remember mapping from row index -> student key for later overlays
    row = 2
    row_key_map = {}
    for reg in regs:
        name = reg.student or reg.lrn
        key = (reg.lrn or reg.student).strip().lower()
        row_key_map[row] = key
        ws.cell(row=row, column=1, value=name)
        for d in range(1, today.day + 1):
            cell = ws.cell(row=row, column=2 + d)
            has_flag = (key in att_flags and d in att_flags[key])
            if has_flag:
                # mark present; we'll overlay triangles later for AM/PM
                cell.value = 'P'
                cell.font = Font(bold=True, color='FF00A050')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.value = 'X'
                cell.font = Font(color='FF000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    # Try to add triangular image overlays for present cells (P) to create a half-cell shading.
    # This is optional: if Pillow isn't installed we'll skip image overlays and keep P/X markers.
    try:
        from PIL import Image as PILImage, ImageDraw
        from openpyxl.drawing.image import Image as OpenpyxlImage

        # Helper to compute approximate pixel sizes (heuristic)
        from openpyxl.utils import column_index_from_string

        def get_column_width_pixels(ws, col_letter: str) -> int:
            default_width = getattr(ws.sheet_format, 'defaultColWidth', None) or 8.43
            cd = ws.column_dimensions.get(col_letter)
            width = getattr(cd, 'width', None) or default_width
            try:
                px = int(float(width) * 7 + 5)
            except Exception:
                px = int(default_width * 7 + 5)
            return max(px, 20)

        def get_row_height_pixels(ws, row_idx: int) -> int:
            default_height = getattr(ws.sheet_format, 'defaultRowHeight', None) or 15
            rd = ws.row_dimensions.get(row_idx)
            height = getattr(rd, 'height', None) or default_height
            try:
                px = int(float(height) * 96.0 / 72.0)
            except Exception:
                px = int(default_height * 96.0 / 72.0)
            return max(px, 12)

        # For each data cell, create a triangle PNG sized to the approximate cell pixels for AM/PM flags
        max_row = row - 1
        for r_idx in range(2, max_row + 1):
            # Determine student key for this row (skip if not present)
            key = row_key_map.get(r_idx)
            if not key:
                continue
            for d in range(1, today.day + 1):
                col_idx = 2 + d
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                try:
                    # flags from att_flags mapping
                    flags = 0
                    if key in att_flags and d in att_flags[key]:
                        flags = att_flags[key][d]
                    if not flags:
                        continue

                    cell_w = get_column_width_pixels(ws, col_letter)
                    cell_h = get_row_height_pixels(ws, r_idx)

                    img_w, img_h = max(4, cell_w), max(4, cell_h)
                    pil_img = PILImage.new('RGBA', (img_w, img_h), (255, 255, 255, 0))
                    draw = ImageDraw.Draw(pil_img)

                    # AM triangle -> top-right (green)
                    if flags & AM_FLAG:
                        am_pts = [(img_w, 0), (img_w, img_h), (0, 0)]
                        draw.polygon(am_pts, fill=(67, 160, 71, 255))
                        try:
                            draw.line([am_pts[0], am_pts[1]], fill=(0,0,0,255), width=1)
                            draw.line([am_pts[1], am_pts[2]], fill=(0,0,0,255), width=1)
                            draw.line([am_pts[2], am_pts[0]], fill=(0,0,0,255), width=1)
                        except Exception:
                            pass

                    # PM triangle -> bottom-left (green)
                    if flags & PM_FLAG:
                        pm_pts = [(0, img_h), (img_w, img_h), (0, 0)]
                        draw.polygon(pm_pts, fill=(67, 160, 71, 255))
                        try:
                            draw.line([pm_pts[0], pm_pts[1]], fill=(0,0,0,255), width=1)
                            draw.line([pm_pts[1], pm_pts[2]], fill=(0,0,0,255), width=1)
                            draw.line([pm_pts[2], pm_pts[0]], fill=(0,0,0,255), width=1)
                        except Exception:
                            pass

                    img_bytes = io.BytesIO()
                    pil_img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)

                    ximg = OpenpyxlImage(img_bytes)
                    ximg.width = img_w
                    ximg.height = img_h
                    anchor = f"{col_letter}{r_idx}"
                    ws.add_image(ximg, anchor)
                except Exception:
                    # If anything fails for a cell, continue — we still want the workbook to be returned
                    continue
    except Exception:
        # Pillow not available or unexpected error — skip adding images
        pass

    # ------ Server-side shading for absent 'X' cells ------
    # Use openpyxl PatternFill to shade cells containing 'X' so shading works even if Pillow
    # is not installed in the environment. This is faster and more reliable than image overlays
    # for full-cell coloring.
    try:
        import re
        # only match an exact single-character X-like glyph
        x_re = re.compile(r"^[xX\u2715\u2716\u00D7\u2718\u2717]$")
        from openpyxl.styles import PatternFill as _PatternFill

        # determine fill color (ARGB) - use pale red
        _fill_hex = 'FFD6D6'
        if len(_fill_hex) == 6 and not _fill_hex.upper().startswith('FF'):
            _fill_argb = 'FF' + _fill_hex
        elif len(_fill_hex) == 8 and _fill_hex.upper().startswith('FF'):
            _fill_argb = _fill_hex
        else:
            _fill_argb = 'FF' + _fill_hex

        _fill = _PatternFill(fill_type='solid', start_color=_fill_argb, end_color=_fill_argb)

        max_row = row - 1 if 'row' in locals() else wb.active.max_row
        for r_idx in range(2, max_row + 1):
            for d in range(1, today.day + 1):
                try:
                    col_idx = 2 + d
                    cell = ws.cell(row=r_idx, column=col_idx)
                    v = cell.value
                    if v is None:
                        continue

                    # Skip formula cells entirely - we don't want to match "x" inside formulas
                    if getattr(cell, 'data_type', None) == 'f':
                        continue
                    if isinstance(v, str) and v.startswith('='):
                        continue

                    s = str(v).strip()
                    if not s:
                        continue

                    # require exact match (single glyph) rather than substring search
                    if not x_re.fullmatch(s):
                        continue

                    # Apply a solid PatternFill so Excel will show the pale-red shading
                    try:
                        cell.fill = _fill
                    except Exception:
                        # best-effort: ignore if fill cannot be applied
                        pass
                except Exception:
                    continue
    except Exception:
        # If anything unexpected happens, continue without blocking the export
        pass

    # Save workbook to exports
    out_dir = settings.BASE_DIR / 'exports'
    try:
        os.makedirs(out_dir, exist_ok=True)
    except Exception:
        out_dir = os.path.join(str(settings.BASE_DIR), 'exports')
        os.makedirs(out_dir, exist_ok=True)

    filename = f"attendance_server_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    dest_path = out_dir / filename if hasattr(out_dir, '__truediv__') else os.path.join(out_dir, filename)
    wb.save(dest_path)

    download_url = f"/exports/{filename}"
    return Response({'filename': filename, 'url': download_url}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([AllowAny])
def registrations_grouped(request):
    """Return registrations grouped by sex (Male/Female) and sorted alphabetically by student name."""
    regs = Registration.objects.all()
    males = regs.filter(sex__iexact='male').order_by('student')
    females = regs.filter(sex__iexact='female').order_by('student')

    male_ser = RegistrationSerializer(males, many=True)
    female_ser = RegistrationSerializer(females, many=True)

    return Response({
        'male': male_ser.data,
        'female': female_ser.data,
    })


@api_view(['POST'])
def drop_registration(request, pk):
    """Copy a registration into DroppedRegistration then delete the original.

    This preserves a server-side record of the dropped student so it can be
    restored later without relying on client localStorage.
    """
    try:
        reg = Registration.objects.get(pk=pk)
    except Registration.DoesNotExist:
        return Response({'error': 'Registration not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        # create dropped copy
        dropped = DroppedRegistration.objects.create(
            lrn=reg.lrn,
            student=reg.student,
            sex=reg.sex,
            parent=reg.parent,
            guardian=reg.guardian,
            original_id=reg.id,
        )

        # delete original registration (this will cascade-delete attendances)
        reg.delete()

        ser = DroppedRegistrationSerializer(dropped)
        return Response(ser.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        # Return a helpful message for debugging (dev only)
        return Response({'error': f'Failed to drop registration: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def dropped_list(request):
    try:
        drops = DroppedRegistration.objects.all().order_by('-dropped_at')
        ser = DroppedRegistrationSerializer(drops, many=True)
        return Response(ser.data)
    except Exception as e:
        return Response({'error': f'Failed to fetch dropped list: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def restore_dropped(request, pk):
    """Restore a dropped registration back into the main Registration table.

    This will create a new Registration using the stored data and remove the
    DroppedRegistration record.
    """
    try:
        dropped = DroppedRegistration.objects.get(pk=pk)
    except DroppedRegistration.DoesNotExist:
        return Response({'error': 'Dropped registration not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Prevent duplicate LRN
        if Registration.objects.filter(lrn=dropped.lrn).exists():
            return Response({'error': 'Registration with same LRN already exists'}, status=status.HTTP_400_BAD_REQUEST)

        reg = Registration.objects.create(
            lrn=dropped.lrn,
            student=dropped.student,
            sex=dropped.sex,
            parent=dropped.parent,
            guardian=dropped.guardian,
        )

        # delete dropped record
        dropped.delete()

        ser = RegistrationSerializer(reg)
        return Response(ser.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': f'Failed to restore dropped registration: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_dropped(request, pk):
    try:
        dropped = DroppedRegistration.objects.get(pk=pk)
    except DroppedRegistration.DoesNotExist:
        return Response({'error': 'Dropped registration not found'}, status=status.HTTP_404_NOT_FOUND)
    try:
        dropped.delete()
        return Response({'message': 'Dropped registration deleted'}, status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        return Response({'error': f'Failed to delete dropped registration: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def generate_half_triangle_excel(request):
    """Generate an Excel file with a half-cell triangular shading by inserting a small PNG triangle overlay.

    This implementation creates a transparent PNG with a filled right-triangle using Pillow and
    inserts it into cell A1. It returns the workbook as an attachment. Pillow is required on the
    server (pip install pillow).
    """
    if openpyxl is None:
        return Response({'error': 'openpyxl is not available on the server'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Import Pillow at runtime and provide a clear error if missing
    try:
        from PIL import Image as PILImage, ImageDraw
    except Exception:
        return Response({'error': 'Pillow (PIL) is required to generate PNG overlays. Please install with pip install pillow'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception:
        return Response({'error': 'openpyxl.drawing.image.Image unavailable'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TriangleDemo'

    # Query params: target cell (e.g. A1), corner (tl,tr,bl,br), color (hex RGB, default green)
    target = (request.GET.get('cell') or 'A1').upper()
    corner = (request.GET.get('corner') or 'tl').lower()
    color_hex = (request.GET.get('color') or '43A047').lstrip('#')

    # Helper: parse hex color to rgba tuple
    def hex_to_rgba(hx: str, alpha: int = 255):
        try:
            if len(hx) == 3:
                hx = ''.join([c*2 for c in hx])
            r = int(hx[0:2], 16)
            g = int(hx[2:4], 16)
            b = int(hx[4:6], 16)
            return (r, g, b, alpha)
        except Exception:
            # fallback to green
            return (67, 160, 71, 255)

    rgba = hex_to_rgba(color_hex, 255)

    # Determine column letter and row number from target (simple parse)
    import re
    m = re.match(r"^([A-Z]+)(\d+)$", target)
    if not m:
        cell_col = 'A'
        cell_row = 1
    else:
        cell_col = m.group(1)
        cell_row = int(m.group(2))

    # Compute approximate pixel size for the cell using column width and row height
    from openpyxl.utils import column_index_from_string

    def get_column_width_pixels(ws, col_letter: str) -> int:
        # Common heuristic: pixels = int(width * 7 + 5)
        default_width = getattr(ws.sheet_format, 'defaultColWidth', None) or 8.43
        cd = ws.column_dimensions.get(col_letter)
        width = getattr(cd, 'width', None) or default_width
        try:
            px = int(float(width) * 7 + 5)
        except Exception:
            px = int(default_width * 7 + 5)
        return max(px, 20)

    def get_row_height_pixels(ws, row_idx: int) -> int:
        # row height is in points; pixels = points * 96/72
        default_height = getattr(ws.sheet_format, 'defaultRowHeight', None) or 15
        rd = ws.row_dimensions.get(row_idx)
        height = getattr(rd, 'height', None) or default_height
        try:
            px = int(float(height) * 96.0 / 72.0)
        except Exception:
            px = int(default_height * 96.0 / 72.0)
        return max(px, 12)

    cell_w = get_column_width_pixels(ws, cell_col)
    cell_h = get_row_height_pixels(ws, cell_row)

    # If the worksheet doesn't have explicit column/row sizes, set them to reasonable values so Excel will render similarly
    try:
        cd = ws.column_dimensions[cell_col]
        if not getattr(cd, 'width', None):
            # reverse heuristic: set width so pixel approximation matches
            cd.width = round((cell_w - 5) / 7, 2)
    except Exception:
        pass
    try:
        rd = ws.row_dimensions[cell_row]
        if not getattr(rd, 'height', None):
            rd.height = round(cell_h * 72.0 / 96.0, 1)
    except Exception:
        pass

    # Create PNG sized to the cell in pixels
    img_w, img_h = max(4, cell_w), max(4, cell_h)
    pil_img = PILImage.new('RGBA', (img_w, img_h), (255, 255, 255, 0))
    draw = ImageDraw.Draw(pil_img)

    # Build triangle points depending on corner
    if corner == 'tl':
        pts = [(0, 0), (0, img_h), (img_w, 0)]
    elif corner == 'tr':
        pts = [(img_w, 0), (img_w, img_h), (0, 0)]
    elif corner == 'bl':
        pts = [(0, img_h), (img_w, img_h), (0, 0)]
    elif corner == 'br':
        pts = [(img_w, img_h), (0, img_h), (img_w, 0)]
    else:
        pts = [(0, 0), (0, img_h), (img_w, 0)]

    # Draw filled triangle
    draw.polygon(pts, fill=rgba)

    # Optionally draw a thin diagonal border for clarity
    border_color = (0, 0, 0, 255)
    try:
        draw.line([pts[0], pts[1]], fill=border_color, width=1)
        draw.line([pts[1], pts[2]], fill=border_color, width=1)
        draw.line([pts[2], pts[0]], fill=border_color, width=1)
    except Exception:
        pass

    # Save PNG to BytesIO and attach via openpyxl Image
    img_bytes = io.BytesIO()
    pil_img.save(img_bytes, format='PNG')
    img_bytes.seek(0)

    ximg = OpenpyxlImage(img_bytes)
    ximg.width = img_w
    ximg.height = img_h

    # Add image anchored to the requested cell
    anchor = f"{cell_col}{cell_row}"
    ws.add_image(ximg, anchor)

    # Optional label for debugging — only add when a query flag debug=1 is present
    if request.GET.get('debug') == '1':
        ws.cell(row=cell_row, column=column_index_from_string(cell_col) + 1, value=f"{anchor} {corner} {cell_w}x{cell_h}px")

    # Save workbook to BytesIO and return as FileResponse
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return FileResponse(out, as_attachment=True, filename='half_triangle.xlsx')


@api_view(['GET'])
def shade_x_cells_excel(request):
    """Produce an Excel file where only cells containing an 'X' (or X-like glyph)
    receive a pale-red full-cell shading applied via a PNG overlay.

    Query params:
    - sheet: optional sheet name to restrict processing (default: all sheets)
    - color: optional hex RGB for shading (default: FFC0CB-like pale red 'FFD6D6')
    - debug=1: write small debug notes next to first detected cell
    """
    if openpyxl is None:
        return Response({'error': 'openpyxl is not available on the server'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Pillow required for PNG overlays
    try:
        from PIL import Image as PILImage, ImageDraw
    except Exception:
        return Response({'error': 'Pillow (PIL) is required to generate PNG overlays. Please install with pip install pillow'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception:
        return Response({'error': 'openpyxl.drawing.image.Image unavailable'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Load template if available (public/attendance_template.xlsx) otherwise create blank workbook
    try:
        template_path = settings.BASE_DIR.parent / 'public' / 'attendance_template.xlsx'
        if template_path.exists():
            wb = openpyxl.load_workbook(str(template_path))
        else:
            wb = openpyxl.Workbook()
    except Exception:
        wb = openpyxl.Workbook()

    # color for shading (hex without #)
    color_hex = (request.GET.get('color') or 'FFD6D6').lstrip('#')

    def hex_to_rgba(hx: str, alpha: int = 255):
        try:
            if len(hx) == 3:
                hx = ''.join([c*2 for c in hx])
            r = int(hx[0:2], 16)
            g = int(hx[2:4], 16)
            b = int(hx[4:6], 16)
            return (r, g, b, alpha)
        except Exception:
            return (255, 214, 214, alpha)

    rgba = hex_to_rgba(color_hex, 255)

    # Helpers to compute approximate pixel size for a cell (same heuristics used elsewhere)
    from openpyxl.utils import get_column_letter

    def get_column_width_pixels(ws, col_letter: str) -> int:
        default_width = getattr(ws.sheet_format, 'defaultColWidth', None) or 8.43
        cd = ws.column_dimensions.get(col_letter)
        width = getattr(cd, 'width', None) or default_width
        try:
            px = int(float(width) * 7 + 5)
        except Exception:
            px = int(default_width * 7 + 5)
        return max(px, 20)

    def get_row_height_pixels(ws, row_idx: int) -> int:
        default_height = getattr(ws.sheet_format, 'defaultRowHeight', None) or 15
        rd = ws.row_dimensions.get(row_idx)
        height = getattr(rd, 'height', None) or default_height
        try:
            px = int(float(height) * 96.0 / 72.0)
        except Exception:
            px = int(default_height * 96.0 / 72.0)
        return max(px, 12)

    # Detection helper for X-like glyphs
    import re
    x_re = re.compile(r"^[xX\u2715\u2716\u00D7\u2718\u2717]$")

    def cell_has_x(cell) -> bool:
        try:
            v = cell.value
            if v is None:
                return False
            # skip formula cells (we don't want to match X inside a formula text like COUNTIF(...,"x"))
            if getattr(cell, 'data_type', None) == 'f':
                return False
            if isinstance(v, str) and v.startswith('='):
                return False

            s = str(v).strip()
            if not s:
                return False

            # Match only if the cell's displayed value is exactly an X-like glyph (or single-char X)
            if x_re.fullmatch(s):
                return True
            # Also accept a single-character X (defensive)
            if len(s) == 1 and x_re.match(s):
                return True
            return False
        except Exception:
            return False

    # Determine which sheets to process
    sheet_name = request.GET.get('sheet')
    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else list(wb.worksheets)

    # For debug, remember first found cell
    first_found = None

    for ws in sheets:
        try:
            max_row = ws.max_row
            max_col = ws.max_column
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    try:
                        cell = ws.cell(row=r, column=c)
                        if not cell_has_x(cell):
                            continue

                        # compute pixel dims for this cell
                        col_letter = get_column_letter(c)
                        cell_w = get_column_width_pixels(ws, col_letter)
                        cell_h = get_row_height_pixels(ws, r)

                        # Prefer a direct cell fill so shading works without Pillow/image overlays.
                        try:
                            # Normalize color_hex into ARGB for PatternFill
                            _hex = color_hex or 'FFD6D6'
                            if len(_hex) == 6 and not _hex.upper().startswith('FF'):
                                _argb = 'FF' + _hex
                            elif len(_hex) == 8 and _hex.upper().startswith('FF'):
                                _argb = _hex
                            else:
                                _argb = 'FF' + _hex
                            fill = PatternFill(fill_type='solid', start_color=_argb, end_color=_argb)
                            cell.fill = fill
                            if first_found is None:
                                first_found = (ws.title, f"{col_letter}{r}")
                        except Exception:
                            # fallback: continue if fill cannot be applied
                            continue
                    except Exception:
                        # ignore per-cell errors
                        continue
        except Exception:
            # ignore sheet-level errors and continue
            continue

    # If debug requested, write a small note near the first found cell
    if request.GET.get('debug') == '1' and first_found:
        try:
            sheet_title, addr = first_found
            ws = wb[sheet_title]
            # put debug note one column to the right
            import re as _re
            m = _re.match(r"^([A-Z]+)(\d+)$", addr)
            if m:
                col = openpyxl.utils.column_index_from_string(m.group(1)) + 1
                row = int(m.group(2))
                ws.cell(row=row, column=col, value=f"shaded {addr}")
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return FileResponse(out, as_attachment=True, filename='shaded_x_cells.xlsx')